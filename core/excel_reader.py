"""
This module provides a fix for the KeyError:
"There is no item named 'xl/sharedStrings.xml'" error
that may occur when using the load_workbook function from the openpyxl library

It's important to note that this solution is a temporary workaround
and might not be the most elegant approach.
However, given the current circumstances, it serves the purpose.
"""
from typing import Type

from openpyxl import Workbook
from openpyxl.reader.excel import ExcelReader
from openpyxl.xml.constants import SHARED_STRINGS
from openpyxl.reader.strings import read_string_table, read_rich_text


class ReworkedReader(ExcelReader):
    """
    A customized Excel reader class that inherits from ExcelReader.
    It is used to read Excel workbook data.

    This class overrides the read_strings method to provide
    a fix for the issue related to missing 'xl/sharedStrings.xml'
    item when using openpyxl's load_workbook function.
    """

    def read_strings(self):
        """
        Reads and processes shared strings from the Excel workbook.

        Changed:
            If shared strings data is not found, an empty list is assigned
            to `self.shared_strings`.
        """
        try:
            ct = self.package.find(SHARED_STRINGS)
            reader = read_string_table
            if self.rich_text:
                reader = read_rich_text
            if ct is not None:
                strings_path = ct.PartName[1:]
                with self.archive.open(strings_path, ) as src:
                    self.shared_strings = reader(src)
        except KeyError:
            self.shared_strings = []


def load_workbook(
        filename: str, read_only: bool = False, keep_vba: bool = False,
        data_only: bool = False, keep_links: bool = True,
        rich_text: bool = False) -> Type[Workbook]:
    """
    Loads an Excel workbook from the given filename.

    Changed:
        the Reader class: to address frequent KeyError issues.
    """

    reader = ReworkedReader(filename, read_only, keep_vba,
                            data_only, keep_links, rich_text
                            )
    reader.read()
    return reader.wb
