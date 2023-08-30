"""
Excel Manager Module

This module provides functions to interact
with Google Sheets and Excel workbooks.
It offers functionalities to load data
from Google Sheets, create or update Excel reports,
and find values within Excel sheets.
"""
import datetime
import os

from typing import List

from httplib2 import ServerNotFoundError
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from pandas import DataFrame, ExcelWriter

from google.auth.exceptions import RefreshError
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError

from .excel_reader import load_workbook
from .settings import TOKEN_FILE_PATH, TOKEN_FILE_NAME, SHEET_NAME, \
    SPREADSHEET_ID, SERVICE_PROVISION_ACCOUNTING_DATE, CONTRACTOR_NAME, \
    UNIQUE_PLACEMENT_NUMBER, SERVICE_PROVISION_ACCOUNTING_MONTH, \
    REPORT_FILE_PATH, REPORT_FILE_NAME, WORKSHEET_DATE_NAME, \
    WORKSHEET_MONTH_NAME


def get_spreadsheet(credentials: Credentials) -> dict:
    """ Load data from Google Sheets. """
    spreadsheet_d: dict = {}

    try:
        resource = build('sheets',
                         'v4',
                         credentials=credentials
                         ).spreadsheets()
        spreadsheet_d = resource.values().get(spreadsheetId=SPREADSHEET_ID,
                                              range=SHEET_NAME
                                              ).execute()
    except RefreshError:
        os.remove(os.path.join(TOKEN_FILE_PATH, TOKEN_FILE_NAME))
    except (ServerNotFoundError, HttpError) as e:
        print(e)
        print('Something went wrong. Check your internet connection'
              ' and try again.'
              )

    return spreadsheet_d


def create_new_rows(spreadsheet: dict) -> list | None:
    """ Create rows from Google Sheets data. """
    rows = []

    values = spreadsheet.get('values', [])
    if values:
        try:
            accounting_date_index = values[0].index(
                SERVICE_PROVISION_ACCOUNTING_DATE
            )
            contractor_index = values[0].index(
                CONTRACTOR_NAME
            )
            placement_id_index = values[0].index(
                UNIQUE_PLACEMENT_NUMBER
            )
            accounting_month_index = values[0].index(
                SERVICE_PROVISION_ACCOUNTING_MONTH
            )
        except (IndexError, ValueError):
            print('The table does not meet the required criteria. '
                  'Please correct it and try again.'
                  )
            return None
    else:
        print('The table is empty. Please correct it and try again.')
        return None

    for row in values:
        rows.append(
            (
                row[contractor_index],
                row[placement_id_index],
                row[accounting_month_index],
                row[accounting_date_index],
            )
        )

    return rows[1:]


def open_or_create_report(data: List[tuple]) -> [Workbook, bool]:
    """ Open or create a report file. """
    report_file_path = os.path.join(REPORT_FILE_PATH, REPORT_FILE_NAME)
    try:
        workbook = load_workbook(filename=report_file_path,
                                 read_only=False,
                                 data_only=True
                                 )
        created = False
    except (FileNotFoundError, KeyError):
        date = datetime.date.today().strftime('%d.%m.%Y')
        data_worksheet_1 = [
            [CONTRACTOR_NAME,
             UNIQUE_PLACEMENT_NUMBER,
             date
             ]
        ]
        data_worksheet_2 = [
            [CONTRACTOR_NAME,
             UNIQUE_PLACEMENT_NUMBER,
             date]
        ]

        for row in data:
            data_worksheet_1.append([row[0], row[1], row[2]])
            data_worksheet_2.append([row[0], row[1], row[3]])
        df1 = DataFrame(data_worksheet_1)
        df2 = DataFrame(data_worksheet_2)

        with ExcelWriter(
                path=report_file_path,
                engine='openpyxl'
        ) as writer:
            df1.to_excel(writer,
                         sheet_name=WORKSHEET_MONTH_NAME,
                         index=False,
                         header=False
                         )
            df2.to_excel(writer,
                         sheet_name=WORKSHEET_DATE_NAME,
                         index=False,
                         header=False
                         )

        created = True
        workbook = writer.book

    return [workbook, created]


def find_on_sheet(sheet: Worksheet, target_value) -> int | None:
    """ Find row with required placement number. """
    for rw_n, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if target_value in row:
            return rw_n

    return None


def update_report(workbook: Workbook, rows: List[tuple]) -> bool:
    """ Update the report workbook with new data."""
    worksheet1, worksheet2 = workbook.worksheets
    length_row_worksheet1 = len(worksheet1[1])
    length_row_worksheet2 = len(worksheet2[1])
    date = datetime.datetime.now().strftime("%d.%m.%Y")
    last_exec = worksheet1.cell(row=1, column=length_row_worksheet1).value
    dt = 0
    if date != last_exec:
        worksheet1.cell(row=1, column=length_row_worksheet1 + 1, value=date)
        worksheet2.cell(row=1, column=length_row_worksheet1 + 1, value=date)
    else:
        dt = 1
    for row in rows:
        update_report_worksheet(worksheet1, row, length_row_worksheet1, dt, 2)
        update_report_worksheet(worksheet2, row, length_row_worksheet2, dt, 3)

    workbook.save(
        os.path.join(REPORT_FILE_PATH, REPORT_FILE_NAME)
    )

    return True


def update_report_worksheet(worksheet: Worksheet,
                            row: tuple, len_row: int, dt,
                            index: int) -> None:
    """ Update a report worksheet with new data. """
    row_number = find_on_sheet(worksheet, row[1])
    if row_number:
        for k, cell in enumerate(worksheet[row_number][-2 + dt:1:-1]):
            if not cell.value:
                continue
            elif cell.value != row[index]:
                worksheet.cell(row=row_number,
                               column=len_row + 1 - dt,
                               value=row[index]
                               )
                break
            elif cell.value == row[index]:
                if not dt:
                    worksheet.cell(row=row_number,
                                   column=len_row + 1 - dt,
                                   value=''
                                   )
                break

    else:
        empty_cells = ('',) * (len_row - 2 - dt)
        worksheet.append(row[:2] + empty_cells + (row[index],))
